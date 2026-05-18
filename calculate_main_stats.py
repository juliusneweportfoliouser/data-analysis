"""Create the compact main-statistics workbook from review_honeypot_comparison.

The input workbook is the output of compare_honeypots.py. This script reads
its Summary and MatchedCohort sheets, verifies the metrics used here against
the underlying per-IP table, and writes a chart-free workbook with the same
table layout as HoneypotAnalysisGraphics - Kopie.xlsx.

Acceptance rates in the generated tables use per-IP mean acceptance:
for each source IP, login_ok_sessions / sessions, then average those rates.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


VARIANTS = ("mod_shell", "mod_llm", "van_shell", "van_llm")
GRAPH_VARIANTS = ("mod_shell", "mod_llm", "van_shell")

BANNER_ROWS = (
    ("Automation", "Sessions with automation banner"),
    ("Interactive-GUI", "Sessions with interactive-GUI banner"),
    ("OpenSSH-CLI", "Sessions with OpenSSH-CLI banner"),
    ("other", "Sessions with 'other' SSH banner"),
    ("empty", "Sessions with empty banner"),
    ("non-SSH", "Sessions with non-SSH banner"),
)

REGULAR_ROWS = (
    ("Total connections", "Total connections"),
    ("Unique source IPs", "Unique source IPs"),
    ("Successful logins (sessions)", "Successful logins (sessions)"),
    ("Login success rate", None),
    ("Unique IPs that logged in", "Unique IPs that logged in"),
)

NORMALIZED_ROWS = (
    ("Unique Commands per 100 sessions", "[NORM/auth] Unique commands per 100 auth sessions"),
    ("Unique hashes per 100 sessions", "[NORM/auth] Unique upload binary hashes per 100 auth sessions"),
    ("Average session length [s]", "Session duration mean (s) [login ok]"),
)

MATCHED_ROWS = (
    ("Unique commands", "unique_commands"),
    ("Total commands", "commands"),
    ("Mean sessions per IP", "sessions"),
    ("Upload events", "upload_events"),
    ("Acceptance rate [%]", "acceptance_rate"),
)


def _number(value) -> float:
    return 0.0 if value is None else float(value)


def _mean(values: Iterable[float]) -> float:
    values = list(values)
    return sum(values) / len(values) if values else 0.0


def _close(a: float, b: float, tolerance: float = 1e-9) -> bool:
    return abs(float(a) - float(b)) <= tolerance


def load_summary(ws) -> dict[str, dict[str, float]]:
    variants = [ws.cell(1, col).value for col in range(2, ws.max_column + 1)]
    data: dict[str, dict[str, float]] = {}
    for row in range(2, ws.max_row + 1):
        metric = ws.cell(row, 1).value
        if not metric:
            continue
        data[str(metric)] = {
            str(variant): ws.cell(row, col).value
            for col, variant in enumerate(variants, start=2)
            if variant
        }
    return data


def load_matched_cohort(ws) -> list[dict]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, values)))
    return rows


def per_ip_acceptance(rows: list[dict], variant: str) -> float:
    session_col = f"sessions_{variant}"
    ok_col = f"login_ok_sessions_{variant}"
    rates = [
        _number(row[ok_col]) / _number(row[session_col])
        for row in rows
        if _number(row.get(session_col)) > 0
    ]
    return _mean(rates)


def matched_rows(rows: list[dict], left: str, right: str) -> list[dict]:
    left_ok = f"login_ok_sessions_{left}"
    right_ok = f"login_ok_sessions_{right}"
    return [
        row for row in rows
        if _number(row.get(left_ok)) > 0 and _number(row.get(right_ok)) > 0
    ]


def matched_metric(rows: list[dict], variant: str, metric: str) -> float:
    if metric == "acceptance_rate":
        return _mean(
            _number(row[f"login_ok_sessions_{variant}"]) / _number(row[f"sessions_{variant}"])
            for row in rows
            if _number(row.get(f"sessions_{variant}")) > 0
        )
    return _mean(_number(row[f"{metric}_{variant}"]) for row in rows)


def validate(summary: dict[str, dict[str, float]], rows: list[dict]) -> list[str]:
    messages: list[str] = []
    errors: list[str] = []

    for variant in VARIANTS:
        sessions = sum(_number(row.get(f"sessions_{variant}")) for row in rows)
        ok = sum(_number(row.get(f"login_ok_sessions_{variant}")) for row in rows)
        unique_ips = sum(1 for row in rows if _number(row.get(f"sessions_{variant}")) > 0)
        login_ips = sum(1 for row in rows if _number(row.get(f"login_ok_sessions_{variant}")) > 0)
        per_ip_acceptance_rate = per_ip_acceptance(rows, variant)

        checks = (
            ("Total connections", sessions),
            ("Unique source IPs", unique_ips),
            ("Successful logins (sessions)", ok),
            ("Unique IPs that logged in", login_ips),
            ("Login success rate", per_ip_acceptance_rate),
        )
        for metric, expected in checks:
            actual = summary[metric][variant]
            if not _close(_number(actual), expected):
                errors.append(
                    f"{metric} {variant}: workbook={actual!r}, expected={expected!r}"
                )

    if errors:
        joined = "\n  ".join(errors)
        raise SystemExit(f"input workbook failed validation:\n  {joined}")

    messages.append(
        "validated Summary counts and per-IP mean login success rates "
        "against MatchedCohort aggregates"
    )
    return messages


def put(ws, cell: str, value, number_format: str | None = None, bold: bool = False) -> None:
    ws[cell] = value
    if number_format:
        ws[cell].number_format = number_format
    if bold:
        ws[cell].font = Font(bold=True)


def write_output(output: Path, summary: dict[str, dict[str, float]], rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Banner classification table.
    for col, variant in enumerate(GRAPH_VARIANTS, start=2):
        put(ws, f"{get_column_letter(col)}1", variant, bold=True)
    for row_idx, (label, metric) in enumerate(BANNER_ROWS, start=2):
        put(ws, f"A{row_idx}", label)
        for col, variant in enumerate(GRAPH_VARIANTS, start=2):
            put(ws, f"{get_column_letter(col)}{row_idx}", summary[metric][variant], "0")

    # Regular analysis table.
    put(ws, "A10", "Regular analysis")
    for col, variant in enumerate(VARIANTS, start=2):
        put(ws, f"{get_column_letter(col)}10", variant, bold=True)
    for row_idx, (label, metric) in enumerate(REGULAR_ROWS, start=11):
        put(ws, f"A{row_idx}", label)
        for col, variant in enumerate(VARIANTS, start=2):
            value = per_ip_acceptance(rows, variant) if metric is None else summary[metric][variant]
            fmt = "0.00%" if label == "Login success rate" else "0"
            put(ws, f"{get_column_letter(col)}{row_idx}", value, fmt)

    # Normalized regular-analysis table.
    put(ws, "A18", "Regular analysis")
    for col, variant in enumerate(GRAPH_VARIANTS, start=2):
        put(ws, f"{get_column_letter(col)}18", variant, bold=True)
    for row_idx, (label, metric) in enumerate(NORMALIZED_ROWS, start=19):
        put(ws, f"A{row_idx}", label)
        for col, variant in enumerate(GRAPH_VARIANTS, start=2):
            if label.startswith("Unique hashes") and variant == "mod_llm":
                put(ws, f"{get_column_letter(col)}{row_idx}", "0*")
            else:
                put(ws, f"{get_column_letter(col)}{row_idx}", summary[metric][variant], "0.00")

    # Matched cohort: modified shell vs vanilla shell.
    write_matched_block(
        ws,
        start_row=24,
        rows=matched_rows(rows, "mod_shell", "van_shell"),
        variants=("mod_shell", "van_shell"),
    )

    # Matched cohort: modified shell vs modified LLM.
    write_matched_block(
        ws,
        start_row=32,
        rows=matched_rows(rows, "mod_shell", "mod_llm"),
        variants=("mod_shell", "mod_llm"),
    )

    ws.column_dimensions["A"].width = 38
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 13

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def write_matched_block(ws, start_row: int, rows: list[dict], variants: tuple[str, str]) -> None:
    put(ws, f"A{start_row}", "Matched cohort analysis")
    for col, variant in enumerate(variants, start=2):
        put(ws, f"{get_column_letter(col)}{start_row}", variant, bold=True)

    for offset, (label, metric) in enumerate(MATCHED_ROWS, start=1):
        row_idx = start_row + offset
        put(ws, f"A{row_idx}", label)
        for col, variant in enumerate(variants, start=2):
            fmt = "0.00%" if metric == "acceptance_rate" else "0.00"
            put(ws, f"{get_column_letter(col)}{row_idx}", matched_metric(rows, variant, metric), fmt)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create the compact main-statistics workbook from review_honeypot_comparison.xlsx."
    )
    parser.add_argument(
        "input",
        type=Path,
        help="Path to review_honeypot_comparison.xlsx generated by compare_honeypots.py.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output workbook path. Default: <input-dir>/HoneypotAnalysisMainStats.xlsx",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input
    output_path = args.output or (input_path.parent / "HoneypotAnalysisMainStats.xlsx")

    wb = load_workbook(input_path, data_only=True, read_only=True)
    required = {"Summary", "MatchedCohort"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise SystemExit(f"input workbook missing required sheets: {', '.join(sorted(missing))}")

    summary = load_summary(wb["Summary"])
    rows = load_matched_cohort(wb["MatchedCohort"])
    for message in validate(summary, rows):
        print(message)

    write_output(output_path, summary, rows)
    print(
        "regular Login success rate output uses per-IP mean acceptance over "
        "all source IPs with sessions"
    )
    print(f"wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
