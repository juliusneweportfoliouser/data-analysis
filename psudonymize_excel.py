"""
Pseudonymize sensitive values in an Excel workbook.

The script creates a copied workbook and modifies only that copy. It reuses
the same salted aliasing implementation as ``pseudonymize_logs.py`` so IPs,
session IDs, HASSH values, UUIDs, and Canarytoken secrets can remain joinable
with pseudonymized Cowrie logs when the same ``--salt-file`` is used.

Usage examples:

    python psudonymize_excel.py --src review_honeypot_comparison.xlsx
    python psudonymize_excel.py --src review_honeypot_comparison.xlsx \
        --dst review_honeypot_comparison_pseudonymized.xlsx
    python psudonymize_excel.py --src review_honeypot_comparison.xlsx \
        --field-only --no-mapping

The reverse mapping is sensitive. Keep it private because it can recover the
original values.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
from collections import Counter
from pathlib import Path
from typing import Callable

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:  # pragma: no cover - exercised only in missing envs
    load_workbook = None
    MergedCell = ()  # type: ignore[assignment]

from pseudonymize_logs import (
    FINGERPRINT_RE,
    HASSH_RE,
    IP_FIELDS,
    IPV4_RE,
    IPV6_RE,
    SESSION_FIELDS,
    SESSION_RE,
    TEXT_FIELDS,
    UUID_RE,
    Pseudonymizer,
    load_or_create_salt,
)


SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
HEADER_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")

EXTRA_IP_HEADERS = {
    "source_ip",
    "source_ips",
    "src_ips",
    "destination_ip",
    "dest_ip",
    "remote_ip",
    "client_ip",
    "triggering_ip",
    "ip_list",
}
EXTRA_SESSION_HEADERS = {
    "first_login_session",
    "transport_id",
}
EXTRA_TOKEN_ID_HEADERS = {
    "token",
    "canarytoken",
    "canarytoken_id",
}
EXTRA_TEXT_HEADERS = {
    "command",
    "commands",
    "client_banner",
    "client_version",
}


def normalize_header(value) -> str:
    """Normalize an Excel column header to a lower-snake-ish token."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = HEADER_NORMALIZE_RE.sub("_", text)
    return re.sub(r"_+", "_", text).strip("_")


def default_output_path(src: Path) -> Path:
    return src.with_name(f"{src.stem}_pseudonymized{src.suffix}")


def default_mapping_path(dst: Path) -> Path:
    return dst.with_name(f"{dst.stem}_pseudonym_mapping.json")


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def fullmatch_ip(value: str) -> bool:
    return bool(IPV4_RE.fullmatch(value) or IPV6_RE.fullmatch(value))


def classify_header(header: str) -> str | None:
    """Return the pseudonymization category implied by a column header."""
    if not header:
        return None
    if header in IP_FIELDS or header in EXTRA_IP_HEADERS:
        return "ip"
    if header in SESSION_FIELDS or header in EXTRA_SESSION_HEADERS:
        return "session"
    if header == "sensor":
        return "sensor"
    if header == "uuid":
        return "uuid"
    if header == "fingerprint":
        return "fingerprint"
    if header == "hassh":
        return "hassh"
    if header == "key":
        return "ssh_key"
    if header == "token_id" or header in EXTRA_TOKEN_ID_HEADERS:
        return "token_id"
    if header in {"token_auth", "auth_token"}:
        return "token_auth"
    if header in TEXT_FIELDS or header in EXTRA_TEXT_HEADERS:
        return "text"
    return None


def pseudonymize_by_category(value: str, category: str, ps: Pseudonymizer) -> str:
    """Apply exact field-level pseudonymization for one cell."""
    if category == "ip":
        if fullmatch_ip(value):
            return ps.ip(value)
        return ps.scrub_text(value)
    if category == "session":
        return ps.session(value)
    if category == "sensor":
        return ps.sensor(value)
    if category == "uuid":
        if UUID_RE.fullmatch(value):
            return ps.uuid(value)
        return ps.scrub_text(value)
    if category == "fingerprint":
        if FINGERPRINT_RE.fullmatch(value):
            return ps.fingerprint(value)
        return ps.scrub_text(value)
    if category == "hassh":
        if HASSH_RE.fullmatch(value):
            return ps.hassh(value)
        return ps.scrub_text(value)
    if category == "ssh_key":
        return ps.ssh_key(value)
    if category == "token_id":
        return ps.token_id(value)
    if category == "token_auth":
        return ps.token_auth(value)
    if category == "text":
        return ps.scrub_text(value)
    return value


def category_replacer(category: str, ps: Pseudonymizer) -> Callable[[str], str]:
    return lambda value: pseudonymize_by_category(value, category, ps)


def replace_known_sensitive_values(text: str, ps: Pseudonymizer) -> str:
    """
    Replace non-regex aliases already learned from exact sensitive columns.

    Token IDs and token auth values are not reliably recognizable by a generic
    regex, so this catches them when they also appear embedded in free text.
    """
    for category in ("token_auth", "token_id", "ssh_key", "sensor"):
        for original, alias in ps._maps.get(category, {}).items():  # noqa: SLF001
            if original in text:
                text = text.replace(original, alias)
    return text


def seed_sensitive_column_values(ws, header_row: int, categories: dict[int, str], ps: Pseudonymizer) -> None:
    """Populate deterministic aliases before the edit pass starts."""
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            category = categories.get(cell.column)
            if category is None or not isinstance(cell.value, str) or is_formula(cell.value):
                continue
            pseudonymize_by_category(cell.value, category, ps)


def pseudonymize_workbook(
    path: Path,
    ps: Pseudonymizer,
    *,
    header_row: int,
    scrub_all_text: bool,
) -> dict[str, Counter]:
    """Pseudonymize workbook cells in place and return per-sheet counters."""
    wb = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    report: dict[str, Counter] = {}

    for ws in wb.worksheets:
        headers = {
            cell.column: normalize_header(cell.value)
            for cell in ws[header_row]
            if cell.value is not None
        }
        categories = {
            column: category
            for column, header in headers.items()
            if (category := classify_header(header)) is not None
        }

        seed_sensitive_column_values(ws, header_row, categories, ps)

        counters: Counter = Counter()
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.row == header_row:
                    continue
                original = cell.value
                if not isinstance(original, str) or is_formula(original):
                    continue

                category = categories.get(cell.column)
                if category is not None:
                    updated = pseudonymize_by_category(original, category, ps)
                    counters[f"field:{category}"] += int(updated != original)
                elif scrub_all_text:
                    updated = ps.scrub_text(original)
                    updated = replace_known_sensitive_values(updated, ps)
                    counters["embedded_text"] += int(updated != original)
                else:
                    continue

                if updated != original:
                    cell.value = updated

        report[ws.title] = counters

    wb.save(path)
    return report


def write_mapping(path: Path, ps: Pseudonymizer) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(ps.inverse_mapping(), f, indent=2, sort_keys=True)
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pseudonymize an Excel workbook copy.")
    parser.add_argument(
        "--src",
        required=True,
        type=Path,
        help="Source .xlsx/.xlsm workbook. This file is never modified.",
    )
    parser.add_argument(
        "--dst",
        type=Path,
        help="Output workbook path (default: <src>_pseudonymized.xlsx).",
    )
    parser.add_argument(
        "--salt-file",
        type=Path,
        default=Path(__file__).with_name(".pseudo_salt"),
        help="Salt file shared with pseudonymize_logs.py for stable aliases.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="One-based row number containing column headers (default: 1).",
    )
    parser.add_argument(
        "--field-only",
        action="store_true",
        help="Only pseudonymize recognized sensitive columns. By default, "
             "other text cells are also scrubbed for embedded IPs, sessions, "
             "UUIDs, fingerprints, HASSH values, and already-seen tokens.",
    )
    parser.add_argument(
        "--mapping-out",
        type=Path,
        help="Reverse mapping path (default: <dst>_pseudonym_mapping.json).",
    )
    parser.add_argument(
        "--no-mapping",
        action="store_true",
        help="Do not write a reverse mapping file.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow replacing an existing output workbook.",
    )
    return parser.parse_args()


def main() -> int:
    if load_workbook is None:
        print("error: openpyxl is required; install dependencies from requirements.txt", file=sys.stderr)
        return 2

    args = parse_args()
    src = args.src
    dst = args.dst or default_output_path(src)

    if not src.is_file():
        print(f"error: source workbook not found: {src}", file=sys.stderr)
        return 2
    if src.suffix.lower() not in SUPPORTED_SUFFIXES:
        print("error: source workbook must be .xlsx or .xlsm", file=sys.stderr)
        return 2
    if args.header_row < 1:
        print("error: --header-row must be >= 1", file=sys.stderr)
        return 2
    if src.resolve() == dst.resolve():
        print("error: --dst must differ from --src", file=sys.stderr)
        return 2
    if dst.exists() and not args.overwrite:
        print(f"error: output already exists: {dst} (use --overwrite)", file=sys.stderr)
        return 2

    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)

    salt = load_or_create_salt(args.salt_file)
    ps = Pseudonymizer(salt)
    report = pseudonymize_workbook(
        dst,
        ps,
        header_row=args.header_row,
        scrub_all_text=not args.field_only,
    )

    if not args.no_mapping:
        write_mapping(args.mapping_out or default_mapping_path(dst), ps)

    changed_total = 0
    for sheet, counters in report.items():
        changed = sum(counters.values())
        changed_total += changed
        if changed:
            details = ", ".join(f"{key}={value}" for key, value in sorted(counters.items()) if value)
            print(f"{sheet}: {changed} cells changed ({details})")

    print(f"done: wrote {dst} ({changed_total} cells changed)")
    if not args.no_mapping:
        print(f"reverse mapping: {args.mapping_out or default_mapping_path(dst)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
